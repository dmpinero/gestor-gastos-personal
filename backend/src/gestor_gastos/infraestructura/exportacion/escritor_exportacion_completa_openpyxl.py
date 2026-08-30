import io

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from gestor_gastos.dominio.exportacion.valores import DatosCompletos


def _texto(valor: str | None) -> str | None:
    return ILLEGAL_CHARACTERS_RE.sub("", valor) if valor is not None else None


class EscritorExportacionCompletaOpenpyxl:
    """Adaptador de EscritorExportacionCompleta que genera el Excel con openpyxl.

    Un libro con una hoja por tabla, cada una con fila de cabecera y los
    campos de la entidad correspondiente (incluido su id, para trazabilidad).
    """

    def escribir(self, datos: DatosCompletos) -> bytes:
        libro = openpyxl.Workbook()
        libro.remove(libro.active)

        hoja = libro.create_sheet("Cuentas")
        hoja.append(["ID", "Número de cuenta", "Alias", "Entidad bancaria", "Moneda", "Titular"])
        for cuenta in datos.cuentas:
            hoja.append(
                [
                    cuenta.id,
                    _texto(cuenta.numero_cuenta),
                    _texto(cuenta.alias),
                    _texto(cuenta.entidad_bancaria),
                    _texto(cuenta.moneda),
                    _texto(cuenta.titular),
                ]
            )

        hoja = libro.create_sheet("Categorías")
        hoja.append(["ID", "Nombre"])
        for categoria in datos.categorias:
            hoja.append([categoria.id, _texto(categoria.nombre)])

        hoja = libro.create_sheet("Subcategorías")
        hoja.append(["ID", "ID categoría", "Nombre"])
        for subcategoria in datos.subcategorias:
            hoja.append([subcategoria.id, subcategoria.categoria_id, _texto(subcategoria.nombre)])

        hoja = libro.create_sheet("Movimientos")
        hoja.append(
            [
                "ID",
                "ID cuenta",
                "ID categoría",
                "ID subcategoría",
                "Fecha valor",
                "Descripción",
                "Comentario",
                "Importe",
                "Saldo",
            ]
        )
        for movimiento in datos.movimientos:
            hoja.append(
                [
                    movimiento.id,
                    movimiento.cuenta_id,
                    movimiento.categoria_id,
                    movimiento.subcategoria_id,
                    movimiento.fecha_valor,
                    _texto(movimiento.descripcion),
                    _texto(movimiento.comentario),
                    movimiento.importe,
                    movimiento.saldo,
                ]
            )

        hoja = libro.create_sheet("Conceptos previstos")
        hoja.append(
            [
                "ID",
                "ID categoría",
                "ID subcategoría",
                "Periodicidad",
                "Importe previsto",
                "Mes de inicio",
            ]
        )
        for concepto in datos.conceptos_previstos:
            hoja.append(
                [
                    concepto.id,
                    concepto.categoria_id,
                    concepto.subcategoria_id,
                    concepto.periodicidad,
                    concepto.importe_previsto,
                    concepto.mes_inicio,
                ]
            )

        hoja = libro.create_sheet("Ajustes mensuales")
        hoja.append(["ID", "ID concepto", "Año", "Mes", "Importe"])
        for ajuste in datos.ajustes:
            hoja.append([ajuste.id, ajuste.concepto_id, ajuste.anio, ajuste.mes, ajuste.importe])

        hoja = libro.create_sheet("Asociaciones")
        hoja.append(
            [
                "ID",
                "ID categoría resumen",
                "ID subcategoría resumen",
                "ID categoría movimiento",
                "ID subcategoría movimiento",
            ]
        )
        for asociacion in datos.asociaciones:
            hoja.append(
                [
                    asociacion.id,
                    asociacion.categoria_resumen_id,
                    asociacion.subcategoria_resumen_id,
                    asociacion.categoria_movimiento_id,
                    asociacion.subcategoria_movimiento_id,
                ]
            )

        buffer = io.BytesIO()
        libro.save(buffer)
        return buffer.getvalue()
