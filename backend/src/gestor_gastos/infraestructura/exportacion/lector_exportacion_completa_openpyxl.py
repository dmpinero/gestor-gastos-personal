import datetime
import io
from decimal import Decimal

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from gestor_gastos.dominio.categoria.entidades import Categoria, Subcategoria
from gestor_gastos.dominio.cuenta.entidades import CuentaBancaria
from gestor_gastos.dominio.exportacion.excepciones import (
    CabeceraExcelNoReconocidaError,
    FilaExcelInvalidaError,
    HojasExcelNoReconocidasError,
)
from gestor_gastos.dominio.exportacion.valores import DatosCompletos
from gestor_gastos.dominio.importacion.excepciones import ExtensionNoSoportadaError
from gestor_gastos.dominio.movimiento.entidades import Movimiento
from gestor_gastos.dominio.prevision.entidades import (
    AjusteMensual,
    AsociacionConcepto,
    ConceptoPrevisto,
)

_EXTENSIONES_SOPORTADAS = {".xlsx"}
_PERIODICIDADES_VALIDAS = {"mensual", "trimestral", "semestral", "anual"}

_CABECERA_CUENTAS = ["ID", "Número de cuenta", "Alias", "Entidad bancaria", "Moneda", "Titular"]
_CABECERA_CATEGORIAS = ["ID", "Nombre"]
_CABECERA_SUBCATEGORIAS = ["ID", "ID categoría", "Nombre"]
_CABECERA_MOVIMIENTOS = [
    "ID",
    "ID cuenta",
    "ID categoría",
    "ID subcategoría",
    "Fecha valor",
    "Descripción",
    "Comentario",
    "Importe",
    "Saldo",
]
_CABECERA_CONCEPTOS_PREVISTOS = [
    "ID",
    "ID categoría",
    "ID subcategoría",
    "Periodicidad",
    "Importe previsto",
    "Mes de inicio",
]
_CABECERA_AJUSTES = ["ID", "ID concepto", "Año", "Mes", "Importe"]
_CABECERA_ASOCIACIONES = [
    "ID",
    "ID categoría resumen",
    "ID subcategoría resumen",
    "ID categoría movimiento",
    "ID subcategoría movimiento",
]

_CABECERAS_ESPERADAS = {
    "Cuentas": _CABECERA_CUENTAS,
    "Categorías": _CABECERA_CATEGORIAS,
    "Subcategorías": _CABECERA_SUBCATEGORIAS,
    "Movimientos": _CABECERA_MOVIMIENTOS,
    "Conceptos previstos": _CABECERA_CONCEPTOS_PREVISTOS,
    "Ajustes mensuales": _CABECERA_AJUSTES,
    "Asociaciones": _CABECERA_ASOCIACIONES,
}


def _entero(valor: object, hoja: str, fila: int, columna: str) -> int:
    if valor is None:
        raise FilaExcelInvalidaError(f"'{columna}' vacío en la hoja '{hoja}', fila {fila}")
    try:
        return int(valor)
    except (TypeError, ValueError) as error:
        raise FilaExcelInvalidaError(
            f"'{columna}' no es un número entero en la hoja '{hoja}', fila {fila}"
        ) from error


def _entero_o_none(valor: object, hoja: str, fila: int, columna: str) -> int | None:
    return None if valor is None else _entero(valor, hoja, fila, columna)


def _texto(valor: object, hoja: str, fila: int, columna: str) -> str:
    texto = str(valor).strip() if valor is not None else ""
    if not texto:
        raise FilaExcelInvalidaError(f"'{columna}' vacío en la hoja '{hoja}', fila {fila}")
    return texto


def _texto_o_none(valor: object) -> str | None:
    if valor is None:
        return None
    texto = str(valor).strip()
    return texto or None


def _decimal(valor: object, hoja: str, fila: int, columna: str) -> Decimal:
    if valor is None:
        raise FilaExcelInvalidaError(f"'{columna}' vacío en la hoja '{hoja}', fila {fila}")
    try:
        return Decimal(str(round(float(valor), 2)))
    except (TypeError, ValueError) as error:
        raise FilaExcelInvalidaError(
            f"'{columna}' no es un número válido en la hoja '{hoja}', fila {fila}"
        ) from error


def _fecha(valor: object, hoja: str, fila: int, columna: str) -> datetime.date:
    if isinstance(valor, datetime.datetime):
        return valor.date()
    if isinstance(valor, datetime.date):
        return valor
    raise FilaExcelInvalidaError(
        f"'{columna}' no es una fecha válida en la hoja '{hoja}', fila {fila}"
    )


def _periodicidad(valor: object, hoja: str, fila: int, columna: str) -> str:
    texto = (_texto_o_none(valor) or "").lower()
    if texto not in _PERIODICIDADES_VALIDAS:
        raise FilaExcelInvalidaError(
            f"'{columna}' no es una periodicidad válida en la hoja '{hoja}', fila {fila}"
        )
    return texto


class LectorExportacionCompletaOpenpyxl:
    """Adaptador de LectorExportacionCompleta para el formato propio generado
    por EscritorExportacionCompletaOpenpyxl (6 hojas, una por tabla, con fila
    de cabecera y columna ID visible)."""

    def leer(self, contenido: bytes, nombre_fichero: str) -> DatosCompletos:
        extension = self._extraer_extension(nombre_fichero)
        if extension not in _EXTENSIONES_SOPORTADAS:
            raise ExtensionNoSoportadaError(
                f"Extensión '{extension}' no soportada; solo se admite .xlsx"
            )

        libro = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        if not all(nombre in libro.sheetnames for nombre in _CABECERAS_ESPERADAS):
            raise HojasExcelNoReconocidasError(
                "El fichero no tiene las 6 hojas esperadas: " + ", ".join(_CABECERAS_ESPERADAS)
            )
        for nombre, cabecera in _CABECERAS_ESPERADAS.items():
            self._validar_cabecera(libro[nombre], nombre, cabecera)

        return DatosCompletos(
            cuentas=self._leer_cuentas(libro["Cuentas"]),
            categorias=self._leer_categorias(libro["Categorías"]),
            subcategorias=self._leer_subcategorias(libro["Subcategorías"]),
            movimientos=self._leer_movimientos(libro["Movimientos"]),
            conceptos_previstos=self._leer_conceptos_previstos(libro["Conceptos previstos"]),
            ajustes=self._leer_ajustes(libro["Ajustes mensuales"]),
            asociaciones=self._leer_asociaciones(libro["Asociaciones"]),
        )

    def _extraer_extension(self, nombre_fichero: str) -> str:
        punto = nombre_fichero.rfind(".")
        return nombre_fichero[punto:].lower() if punto != -1 else ""

    def _validar_cabecera(self, hoja: Worksheet, nombre: str, cabecera_esperada: list[str]) -> None:
        primera_fila = [celda.value for celda in next(hoja.iter_rows(min_row=1, max_row=1), ())]
        if primera_fila != cabecera_esperada:
            raise CabeceraExcelNoReconocidaError(
                f"La hoja '{nombre}' no tiene la cabecera esperada: {', '.join(cabecera_esperada)}"
            )

    def _leer_cuentas(self, hoja: Worksheet) -> list[CuentaBancaria]:
        cuentas: list[CuentaBancaria] = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            if fila[0].value is None:
                break
            cuentas.append(
                CuentaBancaria(
                    id=_entero(fila[0].value, hoja.title, numero_fila, "ID"),
                    numero_cuenta=_texto(
                        fila[1].value, hoja.title, numero_fila, "Número de cuenta"
                    ),
                    alias=_texto_o_none(fila[2].value),
                    entidad_bancaria=_texto_o_none(fila[3].value),
                    moneda=_texto_o_none(fila[4].value),
                    titular=_texto_o_none(fila[5].value),
                )
            )
        return cuentas

    def _leer_categorias(self, hoja: Worksheet) -> list[Categoria]:
        categorias: list[Categoria] = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            if fila[0].value is None:
                break
            categorias.append(
                Categoria(
                    id=_entero(fila[0].value, hoja.title, numero_fila, "ID"),
                    nombre=_texto(fila[1].value, hoja.title, numero_fila, "Nombre"),
                )
            )
        return categorias

    def _leer_subcategorias(self, hoja: Worksheet) -> list[Subcategoria]:
        subcategorias: list[Subcategoria] = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            if fila[0].value is None:
                break
            subcategorias.append(
                Subcategoria(
                    id=_entero(fila[0].value, hoja.title, numero_fila, "ID"),
                    categoria_id=_entero(fila[1].value, hoja.title, numero_fila, "ID categoría"),
                    nombre=_texto(fila[2].value, hoja.title, numero_fila, "Nombre"),
                )
            )
        return subcategorias

    def _leer_movimientos(self, hoja: Worksheet) -> list[Movimiento]:
        movimientos: list[Movimiento] = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            if fila[0].value is None:
                break
            movimientos.append(
                Movimiento(
                    id=_entero(fila[0].value, hoja.title, numero_fila, "ID"),
                    cuenta_id=_entero(fila[1].value, hoja.title, numero_fila, "ID cuenta"),
                    categoria_id=_entero(fila[2].value, hoja.title, numero_fila, "ID categoría"),
                    subcategoria_id=_entero_o_none(
                        fila[3].value, hoja.title, numero_fila, "ID subcategoría"
                    ),
                    fecha_valor=_fecha(fila[4].value, hoja.title, numero_fila, "Fecha valor"),
                    descripcion=_texto(fila[5].value, hoja.title, numero_fila, "Descripción"),
                    comentario=_texto_o_none(fila[6].value),
                    importe=_decimal(fila[7].value, hoja.title, numero_fila, "Importe"),
                    saldo=_decimal(fila[8].value, hoja.title, numero_fila, "Saldo"),
                )
            )
        return movimientos

    def _leer_conceptos_previstos(self, hoja: Worksheet) -> list[ConceptoPrevisto]:
        conceptos: list[ConceptoPrevisto] = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            if fila[0].value is None:
                break
            conceptos.append(
                ConceptoPrevisto(
                    id=_entero(fila[0].value, hoja.title, numero_fila, "ID"),
                    categoria_id=_entero(fila[1].value, hoja.title, numero_fila, "ID categoría"),
                    subcategoria_id=_entero_o_none(
                        fila[2].value, hoja.title, numero_fila, "ID subcategoría"
                    ),
                    periodicidad=_periodicidad(  # type: ignore[arg-type]
                        fila[3].value, hoja.title, numero_fila, "Periodicidad"
                    ),
                    importe_previsto=_decimal(
                        fila[4].value, hoja.title, numero_fila, "Importe previsto"
                    ),
                    mes_inicio=_entero_o_none(
                        fila[5].value, hoja.title, numero_fila, "Mes de inicio"
                    ),
                )
            )
        return conceptos

    def _leer_ajustes(self, hoja: Worksheet) -> list[AjusteMensual]:
        ajustes: list[AjusteMensual] = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            if fila[0].value is None:
                break
            ajustes.append(
                AjusteMensual(
                    id=_entero(fila[0].value, hoja.title, numero_fila, "ID"),
                    concepto_id=_entero(fila[1].value, hoja.title, numero_fila, "ID concepto"),
                    anio=_entero(fila[2].value, hoja.title, numero_fila, "Año"),
                    mes=_entero(fila[3].value, hoja.title, numero_fila, "Mes"),
                    importe=_decimal(fila[4].value, hoja.title, numero_fila, "Importe"),
                )
            )
        return ajustes

    def _leer_asociaciones(self, hoja: Worksheet) -> list[AsociacionConcepto]:
        asociaciones: list[AsociacionConcepto] = []
        for numero_fila, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
            if fila[0].value is None:
                break
            asociaciones.append(
                AsociacionConcepto(
                    id=_entero(fila[0].value, hoja.title, numero_fila, "ID"),
                    categoria_resumen_id=_entero(
                        fila[1].value, hoja.title, numero_fila, "ID categoría resumen"
                    ),
                    subcategoria_resumen_id=_entero_o_none(
                        fila[2].value, hoja.title, numero_fila, "ID subcategoría resumen"
                    ),
                    categoria_movimiento_id=_entero(
                        fila[3].value, hoja.title, numero_fila, "ID categoría movimiento"
                    ),
                    subcategoria_movimiento_id=_entero_o_none(
                        fila[4].value, hoja.title, numero_fila, "ID subcategoría movimiento"
                    ),
                )
            )
        return asociaciones
