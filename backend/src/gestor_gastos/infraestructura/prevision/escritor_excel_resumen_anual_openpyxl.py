import io
from decimal import Decimal

import openpyxl
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook

from gestor_gastos.dominio.prevision.valores import FilaResumenAnual, ResumenAnual

_MESES_CORTOS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]

_PRIMERA_COLUMNA_MES = 5  # E

_FUENTE_PREVISTO = Font(italic=True)
_FUENTE_TOTAL = Font(bold=True)
_RELLENO_AJUSTADO = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")


class EscritorExcelResumenAnualOpenpyxl:
    """Adaptador de EscritorExcelResumenAnual que genera el Excel con openpyxl.

    Formato propio (dos hojas "Gastos"/"Ingresos", columna Año visible y
    columna ID oculta con el id del concepto, para poder reimportar el
    fichero de forma fiable aplicando cada fila a su propio año). Un
    resumen por año del rango exportado, apilados en la misma hoja con una
    fila de totales al final de cada bloque de año.
    """

    def escribir(self, resumenes: list[ResumenAnual]) -> bytes:
        libro = openpyxl.Workbook()
        libro.remove(libro.active)
        self._escribir_hoja(
            libro, "Gastos", [(r.anio, r.filas_gastos, r.totales_gastos) for r in resumenes]
        )
        self._escribir_hoja(
            libro, "Ingresos", [(r.anio, r.filas_ingresos, r.totales_ingresos) for r in resumenes]
        )

        buffer = io.BytesIO()
        libro.save(buffer)
        return buffer.getvalue()

    def _escribir_hoja(
        self,
        libro: Workbook,
        nombre: str,
        bloques_por_anio: list[tuple[int, list[FilaResumenAnual], list[Decimal]]],
    ) -> None:
        hoja = libro.create_sheet(nombre)
        hoja.append(["Año", "ID", "Concepto", "Periodicidad", *_MESES_CORTOS])
        hoja.column_dimensions["B"].hidden = True

        for anio, filas, totales in bloques_por_anio:
            for fila in filas:
                hoja.append(
                    [
                        anio,
                        fila.concepto_id,
                        ILLEGAL_CHARACTERS_RE.sub("", fila.nombre),
                        fila.periodicidad,
                        *[v.importe for v in fila.valores],
                    ]
                )
                fila_excel = hoja.max_row
                for indice, valor in enumerate(fila.valores):
                    celda = hoja.cell(row=fila_excel, column=_PRIMERA_COLUMNA_MES + indice)
                    if valor.origen == "previsto":
                        celda.font = _FUENTE_PREVISTO
                    elif valor.origen == "ajustado":
                        celda.fill = _RELLENO_AJUSTADO

            hoja.append([anio, None, "Total", None, *totales])
            for celda in hoja[hoja.max_row]:
                celda.font = _FUENTE_TOTAL
