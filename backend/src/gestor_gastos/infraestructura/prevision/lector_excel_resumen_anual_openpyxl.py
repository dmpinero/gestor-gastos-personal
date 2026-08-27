import io
from decimal import Decimal

import openpyxl

from gestor_gastos.dominio.importacion.excepciones import ExtensionNoSoportadaError
from gestor_gastos.dominio.prevision.excepciones import (
    FormatoDeIdConceptoInvalidoError,
    HojaExcelNoReconocidaError,
)
from gestor_gastos.dominio.prevision.valores import (
    CeldaResumenAnualExcel,
    DatosResumenAnualExcelLeidos,
)

_HOJAS_ESPERADAS = ("Gastos", "Ingresos")
_PRIMERA_COLUMNA_MES = 5  # E
_EXTENSIONES_SOPORTADAS = {".xlsx"}


class LectorExcelResumenAnualOpenpyxl:
    """Adaptador de LectorExcelResumenAnual para el formato propio generado
    por EscritorExcelResumenAnualOpenpyxl."""

    def leer(self, contenido: bytes, nombre_fichero: str) -> DatosResumenAnualExcelLeidos:
        extension = self._extraer_extension(nombre_fichero)
        if extension not in _EXTENSIONES_SOPORTADAS:
            raise ExtensionNoSoportadaError(
                f"Extensión '{extension}' no soportada; solo se admite .xlsx"
            )

        libro = openpyxl.load_workbook(io.BytesIO(contenido), data_only=True)
        if not all(nombre_hoja in libro.sheetnames for nombre_hoja in _HOJAS_ESPERADAS):
            raise HojaExcelNoReconocidaError(
                "El fichero no tiene las hojas 'Gastos' e 'Ingresos' esperadas"
            )

        celdas: list[CeldaResumenAnualExcel] = []
        for nombre_hoja in _HOJAS_ESPERADAS:
            celdas.extend(self._leer_hoja(libro[nombre_hoja]))

        return DatosResumenAnualExcelLeidos(celdas=celdas)

    def _extraer_extension(self, nombre_fichero: str) -> str:
        punto = nombre_fichero.rfind(".")
        return nombre_fichero[punto:].lower() if punto != -1 else ""

    def _leer_hoja(self, hoja: object) -> list[CeldaResumenAnualExcel]:
        celdas: list[CeldaResumenAnualExcel] = []
        for fila in hoja.iter_rows(min_row=2):
            id_concepto = fila[1].value
            if id_concepto is None:
                continue  # fila de totales u otra fila sin concepto
            try:
                concepto_id = int(id_concepto)
            except (TypeError, ValueError) as error:
                raise FormatoDeIdConceptoInvalidoError(
                    f"La celda de identificador en la hoja '{hoja.title}' no es un número "
                    f"válido ('{id_concepto}'). ¿Has subido el Excel al importador correcto? "
                    "Este panel espera el fichero exportado desde 'Resumen anual', no un "
                    "Excel de conceptos previstos u otro formato."
                ) from error
            anio = int(fila[0].value)
            for indice_mes in range(12):
                valor_celda = fila[_PRIMERA_COLUMNA_MES - 1 + indice_mes].value
                importe = (
                    None if valor_celda is None else Decimal(str(round(float(valor_celda), 2)))
                )
                celdas.append(
                    CeldaResumenAnualExcel(
                        concepto_id=concepto_id, anio=anio, mes=indice_mes + 1, importe=importe
                    )
                )
        return celdas
