import io

import openpyxl
from fastapi.testclient import TestClient
from pytest_bdd import given, parsers, scenarios, then, when

scenarios("exportacion.feature")


@given(parsers.parse('que existe la cuenta "{numero_cuenta}"'), target_fixture="contexto")
@given(parsers.parse('existe la cuenta "{numero_cuenta}"'), target_fixture="contexto")
def existe_la_cuenta(cliente: TestClient, numero_cuenta: str) -> dict:
    cuenta = cliente.post("/api/v1/cuentas", json={"numero_cuenta": numero_cuenta}).json()
    return {"cuenta_id": cuenta["id"]}


@given(
    parsers.parse(
        'existe la categoría "{nombre_categoria}" con la subcategoría "{nombre_subcategoria}"'
    )
)
def existe_la_categoria_con_subcategoria(
    cliente: TestClient, contexto: dict, nombre_categoria: str, nombre_subcategoria: str
) -> None:
    categoria = cliente.post("/api/v1/categorias", json={"nombre": nombre_categoria}).json()
    subcategoria = cliente.post(
        f"/api/v1/categorias/{categoria['id']}/subcategorias", json={"nombre": nombre_subcategoria}
    ).json()
    contexto["categoria_id"] = categoria["id"]
    contexto["subcategoria_id"] = subcategoria["id"]


@given(
    parsers.parse(
        'existe un movimiento en esa cuenta y esa subcategoría con descripción "{descripcion}" '
        'e importe "{importe}"'
    )
)
def existe_un_movimiento(
    cliente: TestClient, contexto: dict, descripcion: str, importe: str
) -> None:
    cliente.post(
        "/api/v1/movimientos",
        json={
            "cuenta_id": contexto["cuenta_id"],
            "categoria_id": contexto["categoria_id"],
            "subcategoria_id": contexto["subcategoria_id"],
            "fecha_valor": "2026-03-15",
            "descripcion": descripcion,
            "importe": importe,
            "saldo": "100.00",
        },
    )


@when("exporto todos los datos", target_fixture="excel_datos_completos")
@given("he exportado todos los datos", target_fixture="excel_datos_completos")
def exporto_todos_los_datos(cliente: TestClient) -> bytes:
    respuesta = cliente.get("/api/v1/exportacion/datos")
    assert respuesta.status_code == 200
    return respuesta.content


@then(
    parsers.parse(
        'el Excel exportado contiene las hojas "{h1}", "{h2}", "{h3}", "{h4}", "{h5}", '
        '"{h6}" y "{h7}"'
    )
)
def el_excel_contiene_las_hojas(
    excel_datos_completos: bytes, h1: str, h2: str, h3: str, h4: str, h5: str, h6: str, h7: str
) -> None:
    libro = openpyxl.load_workbook(io.BytesIO(excel_datos_completos))
    assert libro.sheetnames == [h1, h2, h3, h4, h5, h6, h7]


@then(parsers.parse('la hoja "{hoja}" del Excel exportado contiene la cuenta "{numero_cuenta}"'))
def la_hoja_contiene_la_cuenta(excel_datos_completos: bytes, hoja: str, numero_cuenta: str) -> None:
    libro = openpyxl.load_workbook(io.BytesIO(excel_datos_completos))
    valores_columna_b = [fila[1].value for fila in libro[hoja].iter_rows(min_row=2)]
    assert numero_cuenta in valores_columna_b


@then(
    parsers.parse(
        'la hoja "{hoja}" del Excel exportado contiene un movimiento '
        'con descripción "{descripcion}"'
    )
)
def la_hoja_contiene_un_movimiento(
    excel_datos_completos: bytes, hoja: str, descripcion: str
) -> None:
    libro = openpyxl.load_workbook(io.BytesIO(excel_datos_completos))
    valores_descripcion = [fila[5].value for fila in libro[hoja].iter_rows(min_row=2)]
    assert descripcion in valores_descripcion


@when("importo el backup exportado", target_fixture="resultado_importacion")
def importo_el_backup_exportado(cliente: TestClient, excel_datos_completos: bytes) -> dict:
    respuesta = cliente.post(
        "/api/v1/exportacion/datos/importar",
        files={
            "fichero": (
                "backup-gestor-gastos.xlsx",
                excel_datos_completos,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert respuesta.status_code == 200
    return respuesta.json()


@then(
    parsers.parse(
        "la restauración importa {cuentas:d} cuenta, {categorias:d} categoría y "
        "{movimientos:d} movimiento"
    )
)
def la_restauracion_importa(
    resultado_importacion: dict, cuentas: int, categorias: int, movimientos: int
) -> None:
    assert resultado_importacion["cuentas_importadas"] == cuentas
    assert resultado_importacion["categorias_importadas"] == categorias
    assert resultado_importacion["movimientos_importados"] == movimientos


@then(parsers.parse('solo existe la cuenta "{numero_cuenta}"'))
def solo_existe_la_cuenta(cliente: TestClient, numero_cuenta: str) -> None:
    cuentas = cliente.get("/api/v1/cuentas").json()
    assert [c["numero_cuenta"] for c in cuentas] == [numero_cuenta]
