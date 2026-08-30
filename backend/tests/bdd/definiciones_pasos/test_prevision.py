import io

import openpyxl
from fastapi.testclient import TestClient
from pytest_bdd import given, parsers, scenarios, then, when

scenarios("prevision.feature")


@given(
    parsers.parse(
        'que existe la categoría "{nombre_categoria}" con un concepto previsto '
        'mensual de importe "{importe}"'
    ),
    target_fixture="contexto",
)
def existe_categoria_con_concepto_mensual(
    cliente: TestClient, nombre_categoria: str, importe: str
) -> dict:
    categoria = cliente.post("/api/v1/categorias", json={"nombre": nombre_categoria}).json()
    concepto = cliente.post(
        "/api/v1/previsiones",
        json={
            "categoria_id": categoria["id"],
            "periodicidad": "mensual",
            "importe_previsto": importe,
        },
    ).json()
    return {"categoria_id": categoria["id"], "concepto_id": concepto["id"]}


@given(
    parsers.parse(
        'existe la cuenta "{numero_cuenta}" con un movimiento en esa categoría '
        'en la fecha "{fecha}" e importe "{importe}"'
    )
)
def existe_cuenta_con_movimiento(
    cliente: TestClient, contexto: dict, numero_cuenta: str, fecha: str, importe: str
) -> None:
    cuenta = cliente.post("/api/v1/cuentas", json={"numero_cuenta": numero_cuenta}).json()
    cliente.post(
        "/api/v1/movimientos",
        json={
            "cuenta_id": cuenta["id"],
            "categoria_id": contexto["categoria_id"],
            "fecha_valor": fecha,
            "descripcion": "Movimiento de prueba",
            "importe": importe,
            "saldo": "100.00",
        },
    )


@given(
    parsers.parse(
        'existe la categoría "{nombre_categoria}" con un movimiento de "{importe}" '
        'en la fecha "{fecha}"'
    )
)
def existe_categoria_con_movimiento(
    cliente: TestClient, contexto: dict, nombre_categoria: str, importe: str, fecha: str
) -> None:
    categoria = cliente.post("/api/v1/categorias", json={"nombre": nombre_categoria}).json()
    cuenta = cliente.post(
        "/api/v1/cuentas", json={"numero_cuenta": f"ES00 {nombre_categoria}"}
    ).json()
    cliente.post(
        "/api/v1/movimientos",
        json={
            "cuenta_id": cuenta["id"],
            "categoria_id": categoria["id"],
            "fecha_valor": fecha,
            "descripcion": "Movimiento de prueba",
            "importe": importe,
            "saldo": "100.00",
        },
    )
    contexto["categoria_movimiento_id"] = categoria["id"]


@when(
    parsers.parse(
        'asocio la categoría "{nombre_resumen}" del resumen anual con la categoría '
        '"{nombre_movimiento}" de movimientos'
    )
)
def asocio_categoria_resumen_con_categoria_movimiento(cliente: TestClient, contexto: dict) -> None:
    respuesta = cliente.post(
        "/api/v1/previsiones/asociaciones",
        json={
            "categoria_resumen_id": contexto["categoria_id"],
            "categoria_movimiento_id": contexto["categoria_movimiento_id"],
        },
    )
    assert respuesta.status_code == 201


@given(
    parsers.parse(
        'existe la categoría "{nombre_categoria}" con un movimiento de descripción '
        '"{descripcion}" e importe "{importe}" en la fecha "{fecha}"'
    )
)
def existe_categoria_con_movimiento_con_descripcion(
    cliente: TestClient,
    contexto: dict,
    nombre_categoria: str,
    descripcion: str,
    importe: str,
    fecha: str,
) -> None:
    categoria = cliente.post("/api/v1/categorias", json={"nombre": nombre_categoria}).json()
    cuenta = cliente.post(
        "/api/v1/cuentas", json={"numero_cuenta": f"ES00 {nombre_categoria}"}
    ).json()
    cliente.post(
        "/api/v1/movimientos",
        json={
            "cuenta_id": cuenta["id"],
            "categoria_id": categoria["id"],
            "fecha_valor": fecha,
            "descripcion": descripcion,
            "importe": importe,
            "saldo": "100.00",
        },
    )


@when(
    parsers.parse(
        'asocio la categoría "{nombre_resumen}" del resumen anual con la descripción '
        '"{descripcion}" de movimientos'
    )
)
def asocio_categoria_resumen_con_descripcion_movimiento(
    cliente: TestClient, contexto: dict, descripcion: str
) -> None:
    respuesta = cliente.post(
        "/api/v1/previsiones/asociaciones-descripcion",
        json={
            "categoria_resumen_id": contexto["categoria_id"],
            "descripcion": descripcion,
        },
    )
    assert respuesta.status_code == 201


@given(parsers.parse('se ajusta manualmente el importe del mes {mes:d} de {anio:d} a "{importe}"'))
def se_ajusta_manualmente_el_importe(
    cliente: TestClient, contexto: dict, mes: int, anio: int, importe: str
) -> None:
    respuesta = cliente.put(
        f"/api/v1/previsiones/{contexto['concepto_id']}/ajustes/{anio}/{mes}",
        json={"importe": importe},
    )
    assert respuesta.status_code == 204


@when(parsers.parse("consulto el resumen anual de {anio:d}"), target_fixture="resumen")
def consulto_resumen_anual(cliente: TestClient, anio: int) -> dict:
    return cliente.get(f"/api/v1/previsiones/resumen-anual?anio={anio}").json()


@then(parsers.parse('el concepto muestra el importe real "{importe}" en el mes {mes:d}'))
def el_concepto_muestra_importe_real(resumen: dict, importe: str, mes: int) -> None:
    fila = resumen["filas_gastos"][0]
    valor = next(v for v in fila["valores"] if v["mes"] == mes)
    assert valor["importe"] == importe
    assert valor["origen"] == "real"


@then(parsers.parse('el concepto muestra el importe previsto "{importe}" en el mes {mes:d}'))
def el_concepto_muestra_importe_previsto(resumen: dict, importe: str, mes: int) -> None:
    fila = resumen["filas_gastos"][0]
    valor = next(v for v in fila["valores"] if v["mes"] == mes)
    assert valor["importe"] == importe
    assert valor["origen"] == "previsto"


@then(parsers.parse('el concepto muestra el importe ajustado "{importe}" en el mes {mes:d}'))
def el_concepto_muestra_importe_ajustado(resumen: dict, importe: str, mes: int) -> None:
    fila = resumen["filas_gastos"][0]
    valor = next(v for v in fila["valores"] if v["mes"] == mes)
    assert valor["importe"] == importe
    assert valor["origen"] == "ajustado"


@when(
    parsers.parse("exporto el Excel del resumen anual de {anio:d}"),
    target_fixture="excel_resumen_anual",
)
def exporto_excel_resumen_anual(cliente: TestClient, anio: int) -> bytes:
    respuesta = cliente.get(
        f"/api/v1/previsiones/resumen-anual/exportar?anio_desde={anio}&anio_hasta={anio}"
    )
    assert respuesta.status_code == 200
    return respuesta.content


@when(
    parsers.parse("exporto el Excel del resumen anual de {anio_desde:d} a {anio_hasta:d}"),
    target_fixture="excel_resumen_anual",
)
def exporto_excel_resumen_anual_de_varios_anios(
    cliente: TestClient, anio_desde: int, anio_hasta: int
) -> bytes:
    respuesta = cliente.get(
        f"/api/v1/previsiones/resumen-anual/exportar?anio_desde={anio_desde}&anio_hasta={anio_hasta}"
    )
    assert respuesta.status_code == 200
    return respuesta.content


@when(
    parsers.parse('edito en el Excel exportado el importe del mes {mes:d} a "{importe}"'),
    target_fixture="excel_resumen_anual",
)
def edito_el_excel_exportado(excel_resumen_anual: bytes, mes: int, importe: str) -> bytes:
    libro = openpyxl.load_workbook(io.BytesIO(excel_resumen_anual))
    libro["Gastos"].cell(row=2, column=4 + mes, value=float(importe))
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


@when(
    parsers.parse(
        'edito en el Excel exportado el importe del mes {mes:d} de {anio:d} a "{importe}"'
    ),
    target_fixture="excel_resumen_anual",
)
def edito_el_excel_exportado_de_un_anio(
    excel_resumen_anual: bytes, mes: int, anio: int, importe: str
) -> bytes:
    libro = openpyxl.load_workbook(io.BytesIO(excel_resumen_anual))
    hoja = libro["Gastos"]
    fila_concepto = next(
        fila[0].row
        for fila in hoja.iter_rows(min_row=2)
        if fila[0].value == anio and fila[1].value is not None
    )
    hoja.cell(row=fila_concepto, column=4 + mes, value=float(importe))
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


@when(
    "reimporto el Excel editado",
    target_fixture="resultado_importacion",
)
def reimporto_excel_editado(cliente: TestClient, excel_resumen_anual: bytes) -> dict:
    respuesta = cliente.post(
        "/api/v1/previsiones/resumen-anual/importar",
        files={
            "fichero": (
                "resumen-anual.xlsx",
                excel_resumen_anual,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert respuesta.status_code == 200
    return respuesta.json()


@then(parsers.parse("la importación actualiza {cantidad:d} celda"))
def la_importacion_actualiza_n_celdas(resultado_importacion: dict, cantidad: int) -> None:
    assert resultado_importacion["celdas_actualizadas"] == cantidad


@then(parsers.parse("la importación actualiza {cantidad:d} celdas"))
def la_importacion_actualiza_n_celdas_plural(resultado_importacion: dict, cantidad: int) -> None:
    assert resultado_importacion["celdas_actualizadas"] == cantidad


def _construir_excel_conceptos_previstos(
    categoria: str, subcategoria: str, periodicidad: str, importe: str
) -> bytes:
    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["Categoría", "Subcategoría", "Periodicidad", "Importe previsto"])
    hoja.append([categoria, subcategoria, periodicidad, float(importe)])
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


@when(
    parsers.parse(
        'importo un Excel de conceptos previstos con la fila "{categoria}" / "{subcategoria}" '
        '/ "{periodicidad}" / "{importe}"'
    ),
    target_fixture="resultado_importacion_conceptos",
)
def importo_excel_conceptos_previstos(
    cliente: TestClient, categoria: str, subcategoria: str, periodicidad: str, importe: str
) -> dict:
    contenido = _construir_excel_conceptos_previstos(categoria, subcategoria, periodicidad, importe)
    respuesta = cliente.post(
        "/api/v1/previsiones/importar",
        files={
            "fichero": (
                "conceptos.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert respuesta.status_code == 200
    return respuesta.json()


@given(
    parsers.parse(
        'que ya se importó un Excel de conceptos previstos con la fila "{categoria}" / '
        '"{subcategoria}" / "{periodicidad}" / "{importe}"'
    ),
    target_fixture="excel_conceptos_previstos",
)
def ya_se_importo_excel_conceptos_previstos(
    cliente: TestClient, categoria: str, subcategoria: str, periodicidad: str, importe: str
) -> bytes:
    contenido = _construir_excel_conceptos_previstos(categoria, subcategoria, periodicidad, importe)
    respuesta = cliente.post(
        "/api/v1/previsiones/importar",
        files={
            "fichero": (
                "conceptos.xlsx",
                contenido,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert respuesta.status_code == 200
    return contenido


@when(
    "reimporto el mismo Excel de conceptos previstos",
    target_fixture="resultado_importacion_conceptos",
)
def reimporto_excel_conceptos_previstos(
    cliente: TestClient, excel_conceptos_previstos: bytes
) -> dict:
    respuesta = cliente.post(
        "/api/v1/previsiones/importar",
        files={
            "fichero": (
                "conceptos.xlsx",
                excel_conceptos_previstos,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert respuesta.status_code == 200
    return respuesta.json()


@then(parsers.parse("la importación de conceptos previstos crea {cantidad:d} concepto"))
def la_importacion_crea_n_conceptos(resultado_importacion_conceptos: dict, cantidad: int) -> None:
    assert resultado_importacion_conceptos["conceptos_creados"] == cantidad


@then(
    parsers.parse("la importación de conceptos previstos omite {cantidad:d} concepto por duplicado")
)
def la_importacion_omite_n_conceptos(resultado_importacion_conceptos: dict, cantidad: int) -> None:
    assert resultado_importacion_conceptos["conceptos_omitidos_por_duplicado"] == cantidad


@then(parsers.parse('la importación de conceptos previstos crea la categoría "{nombre}"'))
def la_importacion_crea_la_categoria(resultado_importacion_conceptos: dict, nombre: str) -> None:
    assert nombre in resultado_importacion_conceptos["categorias_creadas"]


@then(parsers.parse('la importación de conceptos previstos crea la subcategoría "{nombre}"'))
def la_importacion_crea_la_subcategoria(resultado_importacion_conceptos: dict, nombre: str) -> None:
    assert nombre in resultado_importacion_conceptos["subcategorias_creadas"]
