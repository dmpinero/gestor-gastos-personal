import io
from decimal import Decimal

import openpyxl

from gestor_gastos.dominio.prevision.valores import FilaResumenAnual, ResumenAnual, ValorMensual
from gestor_gastos.infraestructura.prevision.escritor_excel_resumen_anual_openpyxl import (
    EscritorExcelResumenAnualOpenpyxl,
)


def _valores(importe_por_mes: dict[int, tuple[Decimal, str]]) -> list[ValorMensual]:
    return [
        ValorMensual(mes=mes, importe=importe_por_mes[mes][0], origen=importe_por_mes[mes][1])
        for mes in range(1, 13)
    ]


def test_escribe_las_dos_hojas_gastos_e_ingresos() -> None:
    fila_gasto = FilaResumenAnual(
        concepto_id=1,
        categoria_id=10,
        subcategoria_id=None,
        nombre="Suscripciones",
        periodicidad="mensual",
        valores=_valores({mes: (Decimal("-9.99"), "previsto") for mes in range(1, 13)}),
    )
    resumen = ResumenAnual(
        anio=2026,
        filas_gastos=[fila_gasto],
        filas_ingresos=[],
        totales_gastos=[Decimal("-9.99")] * 12,
        totales_ingresos=[Decimal("0")] * 12,
    )

    contenido = EscritorExcelResumenAnualOpenpyxl().escribir([resumen])
    libro = openpyxl.load_workbook(io.BytesIO(contenido))

    assert libro.sheetnames == ["Gastos", "Ingresos"]


def test_la_columna_anio_va_primero_visible_y_la_id_oculta() -> None:
    fila = FilaResumenAnual(
        concepto_id=42,
        categoria_id=10,
        subcategoria_id=None,
        nombre="Suscripciones",
        periodicidad="mensual",
        valores=_valores({mes: (Decimal("-9.99"), "previsto") for mes in range(1, 13)}),
    )
    resumen = ResumenAnual(
        anio=2026,
        filas_gastos=[fila],
        filas_ingresos=[],
        totales_gastos=[Decimal("-9.99")] * 12,
        totales_ingresos=[Decimal("0")] * 12,
    )

    contenido = EscritorExcelResumenAnualOpenpyxl().escribir([resumen])
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    hoja = libro["Gastos"]

    assert hoja.column_dimensions["B"].hidden is True
    assert hoja["A1"].value == "Año"
    assert hoja["A2"].value == 2026
    assert hoja["B1"].value == "ID"
    assert hoja["B2"].value == 42
    assert hoja["C2"].value == "Suscripciones"
    # Excel no tiene tipo Decimal nativo: se guarda y se relee como número.
    assert hoja["E2"].value == -9.99  # columna E = Ene


def test_la_fila_de_totales_lleva_el_anio_y_va_en_negrita() -> None:
    fila = FilaResumenAnual(
        concepto_id=1,
        categoria_id=10,
        subcategoria_id=None,
        nombre="Suscripciones",
        periodicidad="mensual",
        valores=_valores({mes: (Decimal("-9.99"), "previsto") for mes in range(1, 13)}),
    )
    resumen = ResumenAnual(
        anio=2026,
        filas_gastos=[fila],
        filas_ingresos=[],
        totales_gastos=[Decimal("-9.99")] * 12,
        totales_ingresos=[Decimal("0")] * 12,
    )

    contenido = EscritorExcelResumenAnualOpenpyxl().escribir([resumen])
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    hoja = libro["Gastos"]
    fila_total = hoja[hoja.max_row]

    assert fila_total[0].value == 2026
    assert fila_total[1].value is None
    assert fila_total[2].value == "Total"
    assert fila_total[2].font.bold is True


def test_las_celdas_previstas_van_en_cursiva_y_las_ajustadas_llevan_relleno() -> None:
    valores = _valores({mes: (Decimal("-9.99"), "previsto") for mes in range(1, 13)})
    valores[2] = ValorMensual(mes=3, importe=Decimal("-1.00"), origen="ajustado")
    valores[5] = ValorMensual(mes=6, importe=Decimal("-4.99"), origen="real")
    fila = FilaResumenAnual(
        concepto_id=1,
        categoria_id=10,
        subcategoria_id=None,
        nombre="Suscripciones",
        periodicidad="mensual",
        valores=valores,
    )
    resumen = ResumenAnual(
        anio=2026,
        filas_gastos=[fila],
        filas_ingresos=[],
        totales_gastos=[Decimal("0")] * 12,
        totales_ingresos=[Decimal("0")] * 12,
    )

    contenido = EscritorExcelResumenAnualOpenpyxl().escribir([resumen])
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    hoja = libro["Gastos"]

    celda_previsto = hoja["E2"]  # enero
    celda_ajustado = hoja["G2"]  # marzo
    celda_real = hoja["J2"]  # junio

    assert celda_previsto.font.italic is True
    assert celda_ajustado.fill.start_color.rgb == "00FFF3CD"
    assert not celda_real.font.italic


def test_un_nombre_con_caracteres_de_control_no_rompe_la_escritura() -> None:
    # openpyxl (y el formato XML de .xlsx) no admite ciertos caracteres de
    # control en el contenido de una celda; un nombre de categoría/concepto
    # con uno de ellos no debe tumbar la exportación con IllegalCharacterError.
    fila = FilaResumenAnual(
        concepto_id=1,
        categoria_id=10,
        subcategoria_id=None,
        nombre="Ocio\x00con\x0bcontrol",
        periodicidad="mensual",
        valores=_valores({mes: (Decimal("-9.99"), "previsto") for mes in range(1, 13)}),
    )
    resumen = ResumenAnual(
        anio=2026,
        filas_gastos=[fila],
        filas_ingresos=[],
        totales_gastos=[Decimal("-9.99")] * 12,
        totales_ingresos=[Decimal("0")] * 12,
    )

    contenido = EscritorExcelResumenAnualOpenpyxl().escribir([resumen])
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    hoja = libro["Gastos"]

    assert hoja["C2"].value == "Ocioconcontrol"


def test_varios_anios_se_apilan_en_la_misma_hoja_con_una_fila_de_totales_por_anio() -> None:
    def _resumen(anio: int, concepto_id: int) -> ResumenAnual:
        fila = FilaResumenAnual(
            concepto_id=concepto_id,
            categoria_id=10,
            subcategoria_id=None,
            nombre="Suscripciones",
            periodicidad="mensual",
            valores=_valores({mes: (Decimal("-9.99"), "previsto") for mes in range(1, 13)}),
        )
        return ResumenAnual(
            anio=anio,
            filas_gastos=[fila],
            filas_ingresos=[],
            totales_gastos=[Decimal("-9.99")] * 12,
            totales_ingresos=[Decimal("0")] * 12,
        )

    contenido = EscritorExcelResumenAnualOpenpyxl().escribir([_resumen(2025, 1), _resumen(2026, 1)])
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    hoja = libro["Gastos"]

    # Fila 2: concepto de 2025; fila 3: total de 2025; fila 4: concepto de
    # 2026; fila 5: total de 2026.
    assert hoja["A2"].value == 2025
    assert hoja["C3"].value == "Total"
    assert hoja["A3"].value == 2025
    assert hoja["A4"].value == 2026
    assert hoja["C5"].value == "Total"
    assert hoja["A5"].value == 2026
    assert hoja.max_row == 5
