import datetime
import io
from decimal import Decimal

import openpyxl

from gestor_gastos.dominio.categoria.entidades import Categoria, Subcategoria
from gestor_gastos.dominio.cuenta.entidades import CuentaBancaria
from gestor_gastos.dominio.exportacion.valores import DatosCompletos
from gestor_gastos.dominio.movimiento.entidades import Movimiento
from gestor_gastos.dominio.prevision.entidades import AjusteMensual, ConceptoPrevisto
from gestor_gastos.infraestructura.exportacion.escritor_exportacion_completa_openpyxl import (
    EscritorExportacionCompletaOpenpyxl,
)


def _datos_completos() -> DatosCompletos:
    return DatosCompletos(
        cuentas=[CuentaBancaria(id=1, numero_cuenta="ES00 1234", alias="Principal")],
        categorias=[Categoria(id=10, nombre="Suscripciones")],
        subcategorias=[Subcategoria(id=100, categoria_id=10, nombre="Streaming")],
        movimientos=[
            Movimiento(
                id=1000,
                cuenta_id=1,
                categoria_id=10,
                subcategoria_id=100,
                fecha_valor=datetime.date(2026, 3, 15),
                descripcion="Netflix",
                importe=Decimal("-9.99"),
                saldo=Decimal("100.00"),
            )
        ],
        conceptos_previstos=[
            ConceptoPrevisto(
                id=1,
                categoria_id=10,
                subcategoria_id=100,
                periodicidad="mensual",
                importe_previsto=Decimal("-9.99"),
            )
        ],
        ajustes=[AjusteMensual(id=1, concepto_id=1, anio=2026, mes=3, importe=Decimal("-12.00"))],
    )


def test_escribe_las_seis_hojas() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())
    libro = openpyxl.load_workbook(io.BytesIO(contenido))

    assert libro.sheetnames == [
        "Cuentas",
        "Categorías",
        "Subcategorías",
        "Movimientos",
        "Conceptos previstos",
        "Ajustes mensuales",
    ]


def test_cada_hoja_tiene_cabecera_y_una_fila_por_registro() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())
    libro = openpyxl.load_workbook(io.BytesIO(contenido))

    assert libro["Cuentas"]["A1"].value == "ID"
    assert libro["Cuentas"]["B2"].value == "ES00 1234"
    assert libro["Categorías"]["B2"].value == "Suscripciones"
    assert libro["Subcategorías"]["C2"].value == "Streaming"
    assert libro["Movimientos"]["F2"].value == "Netflix"
    assert libro["Movimientos"]["H2"].value == -9.99
    assert libro["Conceptos previstos"]["D2"].value == "mensual"
    assert libro["Ajustes mensuales"]["E2"].value == -12.0


def test_un_texto_con_caracteres_de_control_no_rompe_la_escritura() -> None:
    datos = _datos_completos()
    datos.categorias[0].nombre = "Ocio\x00con\x0bcontrol"

    contenido = EscritorExportacionCompletaOpenpyxl().escribir(datos)
    libro = openpyxl.load_workbook(io.BytesIO(contenido))

    assert libro["Categorías"]["B2"].value == "Ocioconcontrol"
