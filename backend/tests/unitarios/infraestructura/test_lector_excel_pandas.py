import datetime
from decimal import Decimal
from pathlib import Path

import pytest

from gestor_gastos.dominio.importacion.excepciones import (
    CabeceraNoReconocidaError,
    ExtensionNoSoportadaError,
    FicheroSinMovimientosError,
)
from gestor_gastos.infraestructura.importacion.lector_excel_pandas import (
    LectorExcelPandas,
    _a_texto_o_none,
)

RUTA_FIXTURE = Path(__file__).parent.parent.parent / "fixtures" / "movimientos_ejemplo.xlsx"


def _leer_fixture():
    contenido = RUTA_FIXTURE.read_bytes()
    return LectorExcelPandas().leer(contenido, "movimientos_ejemplo.xlsx")


def test_extrae_numero_de_cuenta_y_titular_de_la_cabecera() -> None:
    datos = _leer_fixture()

    assert datos.cabecera.numero_cuenta == "1234 5678 9012 34567890"
    assert datos.cabecera.titular == "PERSONA EJEMPLO"


def test_extrae_todas_las_filas_de_movimientos() -> None:
    datos = _leer_fixture()

    assert len(datos.filas) == 5


def test_primera_fila_tiene_los_datos_esperados() -> None:
    datos = _leer_fixture()

    primera = datos.filas[0]
    assert primera.fecha_valor == datetime.date(2026, 8, 10)
    assert primera.categoria == "Alimentación"
    assert primera.subcategoria == "Supermercados y alimentación"
    assert primera.importe == Decimal("-25.40")
    assert primera.saldo == Decimal("974.60")
    assert primera.comentario is None


def test_fila_con_comentario_lo_conserva() -> None:
    datos = _leer_fixture()

    fila_con_comentario = next(f for f in datos.filas if f.comentario is not None)
    assert fila_con_comentario.comentario == "Repostaje mensual"


def test_fila_con_importe_positivo_se_lee_igual() -> None:
    datos = _leer_fixture()

    ingreso = next(f for f in datos.filas if f.categoria == "Otros ingresos")
    assert ingreso.importe == Decimal("1200.00")


def test_a_texto_o_none_colapsa_espacios_multiples() -> None:
    assert _a_texto_o_none("Recibo SANITAS S A  DE SEGUROS") == "Recibo SANITAS S A DE SEGUROS"


def test_a_texto_o_none_recorta_espacios_en_los_extremos() -> None:
    assert _a_texto_o_none("  Recibo Ayuntamiento  ") == "Recibo Ayuntamiento"


def test_extension_no_soportada_lanza_error() -> None:
    with pytest.raises(ExtensionNoSoportadaError):
        LectorExcelPandas().leer(b"contenido", "movimientos.csv")


def test_fichero_sin_cabecera_reconocible_lanza_error() -> None:
    import io

    import openpyxl

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.append(["columna_a", "columna_b"])
    hoja.append(["dato_1", "dato_2"])
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(CabeceraNoReconocidaError):
        LectorExcelPandas().leer(buffer.getvalue(), "sin_cabecera.xlsx")


def test_fichero_sin_filas_de_movimientos_lanza_error() -> None:
    import io

    import openpyxl

    libro = openpyxl.Workbook()
    hoja = libro.active
    hoja.cell(row=1, column=3, value="Número de cuenta:")
    hoja.cell(row=1, column=4, value="1234")
    encabezados = [
        "F. VALOR",
        "CATEGORÍA",
        "SUBCATEGORÍA",
        "DESCRIPCIÓN",
        "COMENTARIO",
        "IMPORTE (€)",
        "SALDO (€)",
    ]
    for col, texto in enumerate(encabezados, start=1):
        hoja.cell(row=2, column=col, value=texto)
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(FicheroSinMovimientosError):
        LectorExcelPandas().leer(buffer.getvalue(), "vacio.xlsx")
