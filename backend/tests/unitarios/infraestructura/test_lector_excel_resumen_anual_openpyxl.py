import io
from decimal import Decimal

import openpyxl
import pytest

from gestor_gastos.dominio.importacion.excepciones import ExtensionNoSoportadaError
from gestor_gastos.dominio.prevision.excepciones import HojaExcelNoReconocidaError
from gestor_gastos.infraestructura.prevision.lector_excel_resumen_anual_openpyxl import (
    LectorExcelResumenAnualOpenpyxl,
)


def _libro_con_hojas(*nombres: str) -> bytes:
    libro = openpyxl.Workbook()
    libro.remove(libro.active)
    for nombre in nombres:
        libro.create_sheet(nombre)
    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def _libro_de_ejemplo() -> bytes:
    libro = openpyxl.Workbook()
    libro.remove(libro.active)

    hoja_gastos = libro.create_sheet("Gastos")
    hoja_gastos.append(["ID", "Concepto", "Periodicidad", *[f"M{m}" for m in range(1, 13)]])
    hoja_gastos.append([1, "Suscripciones", "mensual", *[-9.99] * 12])
    hoja_gastos.append([None, "Total", None, *[-9.99] * 12])

    hoja_ingresos = libro.create_sheet("Ingresos")
    hoja_ingresos.append(["ID", "Concepto", "Periodicidad", *[f"M{m}" for m in range(1, 13)]])
    hoja_ingresos.append([2, "Nómina", "mensual", *[2000.00] * 12])
    hoja_ingresos.append([None, "Total", None, *[2000.00] * 12])

    buffer = io.BytesIO()
    libro.save(buffer)
    return buffer.getvalue()


def test_extension_no_soportada_lanza_error() -> None:
    with pytest.raises(ExtensionNoSoportadaError):
        LectorExcelResumenAnualOpenpyxl().leer(b"contenido", "resumen.csv")


def test_hojas_no_reconocidas_lanza_error() -> None:
    contenido = _libro_con_hojas("Hoja1")

    with pytest.raises(HojaExcelNoReconocidaError):
        LectorExcelResumenAnualOpenpyxl().leer(contenido, "resumen-anual-2026.xlsx")


def test_lee_las_celdas_de_gastos_e_ingresos_ignorando_la_fila_de_totales() -> None:
    contenido = _libro_de_ejemplo()

    datos = LectorExcelResumenAnualOpenpyxl().leer(contenido, "resumen-anual-2026.xlsx")

    celdas_concepto_1 = [c for c in datos.celdas if c.concepto_id == 1]
    celdas_concepto_2 = [c for c in datos.celdas if c.concepto_id == 2]
    assert len(celdas_concepto_1) == 12
    assert len(celdas_concepto_2) == 12
    assert all(c.importe == Decimal("-9.99") for c in celdas_concepto_1)
    assert all(c.importe == Decimal("2000.00") for c in celdas_concepto_2)
    assert {c.mes for c in celdas_concepto_1} == set(range(1, 13))


def test_celda_vacia_se_lee_como_importe_none() -> None:
    libro = openpyxl.Workbook()
    libro.remove(libro.active)
    hoja_gastos = libro.create_sheet("Gastos")
    hoja_gastos.append(["ID", "Concepto", "Periodicidad", *[f"M{m}" for m in range(1, 13)]])
    valores = [-9.99] * 12
    valores[0] = None
    hoja_gastos.append([1, "Suscripciones", "mensual", *valores])
    libro.create_sheet("Ingresos").append(
        ["ID", "Concepto", "Periodicidad", *[f"M{m}" for m in range(1, 13)]]
    )
    buffer = io.BytesIO()
    libro.save(buffer)

    datos = LectorExcelResumenAnualOpenpyxl().leer(buffer.getvalue(), "resumen-anual-2026.xlsx")

    celda_enero = next(c for c in datos.celdas if c.mes == 1)
    assert celda_enero.importe is None
