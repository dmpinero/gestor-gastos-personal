import datetime
import io
from decimal import Decimal

import openpyxl
import pytest

from gestor_gastos.dominio.categoria.entidades import Categoria, Subcategoria
from gestor_gastos.dominio.cuenta.entidades import CuentaBancaria
from gestor_gastos.dominio.exportacion.excepciones import (
    CabeceraExcelNoReconocidaError,
    FilaExcelInvalidaError,
    HojasExcelNoReconocidasError,
)
from gestor_gastos.dominio.exportacion.valores import DatosCompletos
from gestor_gastos.dominio.importacion.excepciones import ExtensionNoSoportadaError
from gestor_gastos.dominio.movimiento.entidades import Movimiento
from gestor_gastos.dominio.prevision.entidades import (
    AjusteMensual,
    AsociacionConcepto,
    ConceptoPrevisto,
)
from gestor_gastos.infraestructura.exportacion.escritor_exportacion_completa_openpyxl import (
    EscritorExportacionCompletaOpenpyxl,
)
from gestor_gastos.infraestructura.exportacion.lector_exportacion_completa_openpyxl import (
    LectorExportacionCompletaOpenpyxl,
)


def _datos_completos() -> DatosCompletos:
    return DatosCompletos(
        cuentas=[CuentaBancaria(id=1, numero_cuenta="ES00 1234", alias="Principal")],
        categorias=[Categoria(id=10, nombre="Suscripciones")],
        subcategorias=[Subcategoria(id=100, categoria_id=10, nombre="Streaming")],
        movimientos=[
            Movimiento(
                id=1000,
                cuenta_id=1,
                categoria_id=10,
                subcategoria_id=100,
                fecha_valor=datetime.date(2026, 3, 15),
                descripcion="Netflix",
                importe=Decimal("-9.99"),
                saldo=Decimal("100.00"),
            )
        ],
        conceptos_previstos=[
            ConceptoPrevisto(
                id=1,
                categoria_id=10,
                subcategoria_id=100,
                periodicidad="mensual",
                importe_previsto=Decimal("-9.99"),
            )
        ],
        ajustes=[AjusteMensual(id=1, concepto_id=1, anio=2026, mes=3, importe=Decimal("-12.00"))],
        asociaciones=[
            AsociacionConcepto(
                id=1,
                categoria_resumen_id=10,
                subcategoria_resumen_id=100,
                categoria_movimiento_id=10,
                subcategoria_movimiento_id=100,
            )
        ],
    )


def test_leer_lo_que_escribe_el_escritor_devuelve_los_mismos_datos() -> None:
    originales = _datos_completos()
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(originales)

    leidos = LectorExportacionCompletaOpenpyxl().leer(contenido, "backup-gestor-gastos.xlsx")

    assert leidos == originales


def test_una_extension_no_soportada_lanza_error() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())

    with pytest.raises(ExtensionNoSoportadaError):
        LectorExportacionCompletaOpenpyxl().leer(contenido, "backup.csv")


def test_un_fichero_sin_las_seis_hojas_lanza_error() -> None:
    libro = openpyxl.Workbook()
    libro.active.title = "Cuentas"
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(HojasExcelNoReconocidasError):
        LectorExportacionCompletaOpenpyxl().leer(buffer.getvalue(), "backup.xlsx")


def test_una_cabecera_distinta_de_la_esperada_lanza_error() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    libro["Cuentas"]["A1"] = "Identificador"
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(CabeceraExcelNoReconocidaError):
        LectorExportacionCompletaOpenpyxl().leer(buffer.getvalue(), "backup.xlsx")


def test_un_id_no_numerico_lanza_error() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    libro["Categorías"]["A2"] = "no-es-un-numero"
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(FilaExcelInvalidaError):
        LectorExportacionCompletaOpenpyxl().leer(buffer.getvalue(), "backup.xlsx")


def test_un_nombre_vacio_lanza_error() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    libro["Categorías"]["B2"] = None
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(FilaExcelInvalidaError):
        LectorExportacionCompletaOpenpyxl().leer(buffer.getvalue(), "backup.xlsx")


def test_una_periodicidad_no_reconocida_lanza_error() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    libro["Conceptos previstos"]["D2"] = "quincenal"
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(FilaExcelInvalidaError):
        LectorExportacionCompletaOpenpyxl().leer(buffer.getvalue(), "backup.xlsx")


def test_una_fecha_invalida_lanza_error() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    libro["Movimientos"]["E2"] = "no-es-una-fecha"
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(FilaExcelInvalidaError):
        LectorExportacionCompletaOpenpyxl().leer(buffer.getvalue(), "backup.xlsx")


def test_un_importe_no_numerico_lanza_error() -> None:
    contenido = EscritorExportacionCompletaOpenpyxl().escribir(_datos_completos())
    libro = openpyxl.load_workbook(io.BytesIO(contenido))
    libro["Ajustes mensuales"]["E2"] = "no-es-un-importe"
    buffer = io.BytesIO()
    libro.save(buffer)

    with pytest.raises(FilaExcelInvalidaError):
        LectorExportacionCompletaOpenpyxl().leer(buffer.getvalue(), "backup.xlsx")
